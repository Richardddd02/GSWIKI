#!/usr/bin/env python3
"""Extract 古神编年史 workbook content into a static-site dataset."""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import posixpath
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image

NS = {
    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "v": "urn:schemas-microsoft-com:vml",
    "o": "urn:schemas-microsoft-com:office:office",
    "x": "urn:schemas-microsoft-com:office:excel",
}
R_ID = f"{{{NS['r']}}}id"
R_EMBED = f"{{{NS['r']}}}embed"
O_RELID = f"{{{NS['o']}}}relid"


def clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", "\n").split())


def cell_ref(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def stable_id(*parts: str) -> str:
    raw = "|".join(parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


class WorkbookMedia:
    def __init__(self, source: Path, output_root: Path):
        self.source = source
        self.output_root = output_root
        self.media_dir = output_root / "assets" / "media"
        self.media_dir.mkdir(parents=True, exist_ok=True)
        self.zip = zipfile.ZipFile(source)
        self.sheet_paths = self._sheet_paths()
        self._normal_cache: dict[str, dict[tuple[int, int], list[dict[str, Any]]]] = {}
        self._note_cache: dict[str, dict[tuple[int, int], str]] = {}
        self._dimensions: dict[str, tuple[int, int]] = {}
        self._copied: set[str] = set()

    def close(self) -> None:
        self.zip.close()

    def _rels_path(self, owner_path: str) -> str:
        return posixpath.join(
            posixpath.dirname(owner_path),
            "_rels",
            posixpath.basename(owner_path) + ".rels",
        )

    def _relationships(self, owner_path: str) -> dict[str, str]:
        rel_path = self._rels_path(owner_path)
        if rel_path not in self.zip.namelist():
            return {}
        root = ET.fromstring(self.zip.read(rel_path))
        base = posixpath.dirname(owner_path)
        return {
            rel.attrib["Id"]: posixpath.normpath(posixpath.join(base, rel.attrib["Target"]))
            for rel in root.findall("pr:Relationship", NS)
            if "Target" in rel.attrib
        }

    def _sheet_paths(self) -> dict[str, str]:
        workbook_path = "xl/workbook.xml"
        root = ET.fromstring(self.zip.read(workbook_path))
        rels = self._relationships(workbook_path)
        result: dict[str, str] = {}
        for sheet in root.findall("m:sheets/m:sheet", NS):
            result[sheet.attrib["name"]] = rels[sheet.attrib[R_ID]]
        return result

    def image_dimensions(self, media_path: str) -> tuple[int, int]:
        if media_path not in self._dimensions:
            with Image.open(io.BytesIO(self.zip.read(media_path))) as image:
                self._dimensions[media_path] = image.size
        return self._dimensions[media_path]

    def normal_images(self, sheet_name: str) -> dict[tuple[int, int], list[dict[str, Any]]]:
        if sheet_name in self._normal_cache:
            return self._normal_cache[sheet_name]
        sheet_path = self.sheet_paths[sheet_name]
        root = ET.fromstring(self.zip.read(sheet_path))
        rels = self._relationships(sheet_path)
        drawing = root.find("m:drawing", NS)
        result: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
        if drawing is not None and drawing.attrib.get(R_ID) in rels:
            drawing_path = rels[drawing.attrib[R_ID]]
            drawing_root = ET.fromstring(self.zip.read(drawing_path))
            drawing_rels = self._relationships(drawing_path)
            for anchor in list(drawing_root):
                start = anchor.find("xdr:from", NS)
                blip = anchor.find(".//a:blip", NS)
                if start is None or blip is None:
                    continue
                rid = blip.attrib.get(R_EMBED)
                if rid not in drawing_rels:
                    continue
                row_node = start.find("xdr:row", NS)
                col_node = start.find("xdr:col", NS)
                if row_node is None or col_node is None:
                    continue
                media_path = drawing_rels[rid]
                try:
                    width, height = self.image_dimensions(media_path)
                except Exception:
                    continue
                result[(int(row_node.text) + 1, int(col_node.text) + 1)].append(
                    {"media": media_path, "width": width, "height": height}
                )
        self._normal_cache[sheet_name] = dict(result)
        return self._normal_cache[sheet_name]

    def note_images(self, sheet_name: str) -> dict[tuple[int, int], str]:
        if sheet_name in self._note_cache:
            return self._note_cache[sheet_name]
        sheet_path = self.sheet_paths[sheet_name]
        root = ET.fromstring(self.zip.read(sheet_path))
        rels = self._relationships(sheet_path)
        legacy = root.find("m:legacyDrawing", NS)
        result: dict[tuple[int, int], str] = {}
        if legacy is not None and legacy.attrib.get(R_ID) in rels:
            vml_path = rels[legacy.attrib[R_ID]]
            vml_root = ET.fromstring(self.zip.read(vml_path))
            vml_rels = self._relationships(vml_path)
            for shape in vml_root.findall("v:shape", NS):
                client = shape.find("x:ClientData", NS)
                fill = shape.find("v:fill", NS)
                if client is None or fill is None:
                    continue
                row_node = client.find("x:Row", NS)
                col_node = client.find("x:Column", NS)
                rid = fill.attrib.get(O_RELID)
                if row_node is None or col_node is None or rid not in vml_rels:
                    continue
                result[(int(row_node.text) + 1, int(col_node.text) + 1)] = vml_rels[rid]
        self._note_cache[sheet_name] = result
        return result

    def copy_asset(self, media_path: str | None) -> str:
        if not media_path:
            return ""
        filename = posixpath.basename(media_path)
        destination = self.media_dir / filename
        if media_path not in self._copied:
            destination.write_bytes(self.zip.read(media_path))
            self._copied.add(media_path)
        return f"assets/media/{filename}"

    def icon_for(self, sheet_name: str, row: int, col: int) -> str:
        candidates = self.normal_images(sheet_name).get((row, col), [])
        if not candidates:
            return ""
        icon_candidates = [
            item
            for item in candidates
            if 24 <= item["width"] <= 320
            and 24 <= item["height"] <= 320
            and 0.45 <= item["width"] / item["height"] <= 2.2
        ]
        selected = max(icon_candidates or candidates, key=lambda item: item["width"] * item["height"])
        return self.copy_asset(selected["media"])

    def detail_for(self, sheet_name: str, row: int, col: int) -> str:
        return self.copy_asset(self.note_images(sheet_name).get((row, col)))

    def largest_image(self, sheet_name: str) -> str:
        all_images = [item for items in self.normal_images(sheet_name).values() for item in items]
        if not all_images:
            return ""
        selected = max(all_images, key=lambda item: item["width"] * item["height"])
        return self.copy_asset(selected["media"])


def make_entry(
    *,
    category: str,
    name: str,
    sheet: str,
    row: int,
    col: int,
    subtype: str = "",
    quality: str = "",
    profession: str = "",
    summary: str = "",
    source: str = "",
    usage: str = "",
    image: str = "",
    detail_image: str = "",
    tags: list[str] | None = None,
) -> dict[str, Any]:
    return {
        "id": stable_id(category, sheet, cell_ref(row, col), name),
        "category": category,
        "name": name,
        "sheet": sheet,
        "cell": cell_ref(row, col),
        "subtype": subtype,
        "quality": quality,
        "profession": profession,
        "summary": summary,
        "source": source,
        "usage": usage,
        "image": image,
        "detailImage": detail_image,
        "tags": [tag for tag in (tags or []) if tag],
    }


def extract_named_note_grid(wb, media: WorkbookMedia, sheet: str, quality: str) -> list[dict[str, Any]]:
    ws = wb[sheet]
    entries = []
    for row, col in sorted(media.note_images(sheet)):
        name = clean(ws.cell(row + 1, col).value)
        if not name:
            continue
        entries.append(
            make_entry(
                category="equipment",
                name=name,
                sheet=sheet,
                row=row,
                col=col,
                subtype=f"{quality}装备",
                quality=quality,
                summary=f"{quality}品质装备，详细属性以图鉴记录为准。",
                image=media.icon_for(sheet, row, col),
                detail_image=media.detail_for(sheet, row, col),
                tags=["装备", quality],
            )
        )
    return entries


def extract_sets(wb, media: WorkbookMedia) -> list[dict[str, Any]]:
    entries = []
    numerals = {3: "一", 4: "二", 5: "三"}
    for profession in ["战士", "猎人", "法师", "神官", "佣兵"]:
        sheet = f"{profession}套装"
        ws = wb[sheet]
        for row, col in sorted(media.note_images(sheet)):
            group = numerals.get(row, str(row - 2))
            part = col - 4
            bonus = clean(ws.cell(row, 10).value)
            name = f"{profession}套装 · 第{group}组部件 {part}"
            entries.append(
                make_entry(
                    category="equipment",
                    name=name,
                    sheet=sheet,
                    row=row,
                    col=col,
                    subtype="套装装备",
                    quality="套装",
                    profession=profession,
                    summary=bonus or f"{profession}职业套装部件。",
                    image=media.icon_for(sheet, row, col),
                    detail_image=media.detail_for(sheet, row, col),
                    tags=["装备", "套装", profession, f"第{group}组"],
                )
            )
    return entries


def extract_base_gear(wb, media: WorkbookMedia) -> list[dict[str, Any]]:
    sheet = "基底装备"
    ws = wb[sheet]
    entries = []
    blocks = [(2, 4, 5), (8, 10, 11), (14, 16, 17), (20, 22, 23)]
    for part_col, name_col, attr_col in blocks:
        part = ""
        for row in range(3, 18):
            next_part = clean(ws.cell(row, part_col).value)
            if next_part and next_part != "部位":
                part = next_part
            name = clean(ws.cell(row, name_col).value)
            if not name or not part:
                continue
            attribute = clean(ws.cell(row, attr_col).value)
            entries.append(
                make_entry(
                    category="equipment",
                    name=name,
                    sheet=sheet,
                    row=row,
                    col=name_col,
                    subtype=part,
                    quality="基底",
                    summary=attribute,
                    image=media.icon_for(sheet, row, name_col - 1),
                    tags=["装备", "基底", part],
                )
            )
    return entries


def latest_label(ws, row: int, allowed: set[str]) -> str:
    for current in range(row, 0, -1):
        value = clean(ws.cell(current, 1).value)
        if value in allowed:
            return value
    return ""


def extract_skills(wb, media: WorkbookMedia) -> list[dict[str, Any]]:
    entries = []
    sections = {"基础技能", "被动技能", "主动技能"}
    for profession in ["战士", "猎人", "法师", "神官", "佣兵"]:
        sheet = f"{profession}技能"
        ws = wb[sheet]
        for row, col in sorted(media.note_images(sheet)):
            name = clean(ws.cell(row + 1, col).value)
            if not name:
                continue
            subtype = latest_label(ws, row, sections)
            entries.append(
                make_entry(
                    category="skill",
                    name=name,
                    sheet=sheet,
                    row=row,
                    col=col,
                    subtype=subtype,
                    profession=profession,
                    summary=f"{profession}{subtype}，技能数值与说明见完整档案。",
                    image=media.icon_for(sheet, row, col),
                    detail_image=media.detail_for(sheet, row, col),
                    tags=["技能", profession, subtype],
                )
            )
    return entries


def extract_runes(wb, media: WorkbookMedia) -> list[dict[str, Any]]:
    entries = []
    for quality in ["普通", "精良", "稀有", "传说"]:
        sheet = f"{quality}符文"
        ws = wb[sheet]
        for row, col in sorted(media.note_images(sheet)):
            name = clean(ws.cell(row + 1, col).value)
            if not name:
                continue
            entries.append(
                make_entry(
                    category="rune",
                    name=name,
                    sheet=sheet,
                    row=row,
                    col=col,
                    subtype="符文",
                    quality=quality,
                    summary=f"{quality}品质符文，具体效果见符文档案。",
                    image=media.icon_for(sheet, row, col),
                    detail_image=media.detail_for(sheet, row, col),
                    tags=["符文", quality],
                )
            )
    return entries


def extract_rune_words(wb, media: WorkbookMedia) -> list[dict[str, Any]]:
    sheet = "符文之语"
    ws = wb[sheet]
    professions = {"战士", "猎人", "法师", "神官", "佣兵"}
    entries = []
    for row, col in sorted(media.note_images(sheet)):
        name = clean(ws.cell(row + 1, col).value)
        if not name:
            continue
        profession = latest_label(ws, row, professions)
        entries.append(
            make_entry(
                category="runeword",
                name=name,
                sheet=sheet,
                row=row,
                col=col,
                subtype="符文之语",
                quality="符文之语",
                profession=profession,
                summary=f"{profession}技能对应的三孔符文之语。",
                image=media.icon_for(sheet, row, col),
                detail_image=media.detail_for(sheet, row, col),
                tags=["符文之语", profession],
            )
        )
    return entries


def extract_items(wb, media: WorkbookMedia) -> list[dict[str, Any]]:
    sheet = "道具介绍"
    ws = wb[sheet]
    entries = []
    blocks = [
        (1, 2, 4, 8, range(2, 17)),
        (12, 13, 15, 19, range(2, 17)),
    ]
    for icon_col, name_col, source_col, use_col, rows in blocks:
        for row in rows:
            name = clean(ws.cell(row, name_col).value)
            if not name or name == "名称":
                continue
            source = clean(ws.cell(row, source_col).value)
            usage = clean(ws.cell(row, use_col).value)
            entries.append(
                make_entry(
                    category="item",
                    name=name,
                    sheet=sheet,
                    row=row,
                    col=name_col,
                    subtype="道具",
                    summary=usage or source,
                    source=source,
                    usage=usage,
                    image=media.icon_for(sheet, row, icon_col),
                    tags=["道具"],
                )
            )
    return entries


def build(source: Path, output_root: Path) -> dict[str, Any]:
    media_path = output_root / "assets" / "media"
    if media_path.exists():
        shutil.rmtree(media_path)
    media_path.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(source, data_only=False, read_only=False)
    media = WorkbookMedia(source, output_root)
    try:
        entries: list[dict[str, Any]] = []
        entries.extend(extract_named_note_grid(wb, media, "传说装备", "传说"))
        entries.extend(extract_sets(wb, media))
        entries.extend(extract_named_note_grid(wb, media, "独特装备", "独特"))
        entries.extend(extract_named_note_grid(wb, media, "神话装备", "神话"))
        entries.extend(extract_base_gear(wb, media))
        entries.extend(extract_skills(wb, media))
        entries.extend(extract_runes(wb, media))
        entries.extend(extract_rune_words(wb, media))
        entries.extend(extract_items(wb, media))

        category_order = {"equipment": 0, "skill": 1, "rune": 2, "runeword": 3, "item": 4}
        entries.sort(key=lambda item: (category_order[item["category"]], item["quality"], item["profession"], item["name"]))

        counts: dict[str, int] = defaultdict(int)
        for entry in entries:
            counts[entry["category"]] += 1

        payload = {
            "meta": {
                "title": "古神编年史",
                "game": "放置之路：古神复苏",
                "version": "0.3.2.0",
                "source": source.name,
                "heroImage": media.largest_image("主页"),
                "total": len(entries),
                "counts": dict(counts),
            },
            "entries": entries,
        }
        data_path = output_root / "data.js"
        data_path.write_text(
            "window.CHRONICLE_DATA = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n",
            encoding="utf-8",
        )
        return payload
    finally:
        media.close()
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("--output", type=Path, default=Path(__file__).resolve().parents[1])
    args = parser.parse_args()
    payload = build(args.source.resolve(), args.output.resolve())
    print(json.dumps({"total": payload["meta"]["total"], "counts": payload["meta"]["counts"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
